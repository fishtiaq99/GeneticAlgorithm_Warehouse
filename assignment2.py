"""
Assignment 2: Warehouse Storage Optimization Using Genetic Algorithm
Author: Fatima Ishtiaq
"""

import random
import copy
from itertools import combinations

# ─────────────────────────────────────────────
# 1. WAREHOUSE ZONES
# ─────────────────────────────────────────────
ZONES = {
    "Z1": {"name": "Heavy Item Floor Storage",          "capacity": 120},
    "Z2": {"name": "Standard Rack Storage",             "capacity": 80},
    "Z3": {"name": "Fragile Item Shelf",                "capacity": 80},
    "Z4": {"name": "Temperature Controlled Storage",    "capacity": 80},
    "Z5": {"name": "Hazardous Material Storage",        "capacity": 80},
    "Z6": {"name": "Fast-Moving Product Area",          "capacity": 60},
    "Z7": {"name": "Bulk Dry Storage",                  "capacity": 150},
    "Z8": {"name": "Refrigerated Loading Dock",         "capacity": 100},
}
ZONE_KEYS = list(ZONES.keys())  # ["Z1", ..., "Z8"]

# ─────────────────────────────────────────────
# 2. PRODUCT DATASET
# ─────────────────────────────────────────────
PRODUCTS = [
    # (name,               weight, category,     fragile, hazardous, temp_ctrl, demand)
    ("P1 - Glass Bottles",       20, "Beverage",   True,  False, False, "Medium"),
    ("P2 - Frozen Meat",         30, "Food",       False, False, True,  "High"),
    ("P3 - Cleaning Acid",       10, "Chemical",   False, True,  False, "Low"),
    ("P4 - Rice Bags",           50, "Grocery",    False, False, False, "High"),
    ("P5 - Ceramic Plates",      15, "Kitchenware",True,  False, False, "Medium"),
    ("P6 - Ice Cream",           25, "Food",       False, False, True,  "High"),
    ("P7 - Detergent",           12, "Chemical",   False, True,  False, "Medium"),
    ("P8 - Chips Carton",         8, "Snacks",     False, False, False, "High"),
    ("P9 - Olive Oil Bottles",   18, "Grocery",    True,  False, False, "Medium"),
    ("P10 - Industrial Bleach",  22, "Chemical",   False, True,  False, "Low"),
    ("P11 - Yogurt Cartons",     14, "Food",       False, False, True,  "High"),
    ("P12 - Flour Bags",         45, "Grocery",    False, False, False, "High"),
    ("P13 - Wine Bottles",       16, "Beverage",   True,  False, False, "Medium"),
    ("P14 - Paint Cans",         28, "Chemical",   False, True,  False, "Low"),
    ("P15 - Biscuit Boxes",       6, "Snacks",     False, False, False, "High"),
    ("P16 - Motor Oil",          35, "Chemical",   False, True,  False, "Low"),
    ("P17 - Frozen Fish",        20, "Food",       False, False, True,  "High"),
    ("P18 - Bubble Wrap Rolls",  10, "Packaging",  True,  False, False, "Medium"),
    ("P19 - Wheat Sacks",        40, "Grocery",    False, False, False, "High"),
    ("P20 - Hand Sanitizer",      8, "Chemical",   True,  False, False, "Medium"),
]

NUM_PRODUCTS = len(PRODUCTS)

# ─────────────────────────────────────────────
# 3. CONSTRAINT SETS (indices into PRODUCTS, 0-based)
# ─────────────────────────────────────────────
FRAGILE_IDX      = [i for i, p in enumerate(PRODUCTS) if p[3]]          # must go Z3
HAZARDOUS_IDX    = [i for i, p in enumerate(PRODUCTS) if p[4]]          # must go Z5
TEMP_IDX         = [i for i, p in enumerate(PRODUCTS) if p[5]]          # Z4 or Z8
HIGH_DEMAND_IDX  = [i for i, p in enumerate(PRODUCTS) if p[6] == "High"]# should be Z6
HEAVY_IDX        = [i for i, p in enumerate(PRODUCTS) if p[1] > 40]     # must be Z1
FOOD_IDX         = [i for i, p in enumerate(PRODUCTS) if p[2] == "Food"]
CHEMICAL_IDX     = [i for i, p in enumerate(PRODUCTS) if p[2] == "Chemical"]
# Z8 eligible: temp-controlled AND high-demand
Z8_ELIGIBLE_IDX  = [i for i in TEMP_IDX if PRODUCTS[i][6] == "High"]

# ─────────────────────────────────────────────
# 4. PENALTIES
# ─────────────────────────────────────────────
PENALTY_WEIGHT_OVER   = 2   # per kg over capacity
PENALTY_FRAGILE       = 8   # per product not in Z3
PENALTY_HAZARDOUS     = 10  # per product not in Z5
PENALTY_TEMP          = 9   # per product not in Z4/Z8
PENALTY_DEMAND        = 5   # per high-demand product not in Z6
PENALTY_HEAVY         = 4   # per heavy product not in Z1
PENALTY_COMPAT        = 3   # per separated same-category pair
PENALTY_FOOD_CHEM     = 15  # per food-chemical pair in same zone
PENALTY_Z8_NONELIG    = 12  # per non-eligible product in Z8

# ─────────────────────────────────────────────
# 5. FITNESS FUNCTION
# ─────────────────────────────────────────────
def fitness(chromosome):
    """Lower is better. 0 = fully feasible solution."""
    penalty = 0
    violations = []

    # Build zone → product-index mapping
    zone_products = {z: [] for z in ZONE_KEYS}
    for prod_idx, zone in enumerate(chromosome):
        zone_products[zone].append(prod_idx)

    # 5.1 Zone weight capacity
    for zone, idxs in zone_products.items():
        total_w = sum(PRODUCTS[i][1] for i in idxs)
        cap = ZONES[zone]["capacity"]
        if total_w > cap:
            excess = total_w - cap
            pen = excess * PENALTY_WEIGHT_OVER
            penalty += pen
            violations.append(f"[Weight] {zone} over by {excess} kg → +{pen}")

    # 5.2 Fragile products must be in Z3
    for i in FRAGILE_IDX:
        if chromosome[i] != "Z3":
            penalty += PENALTY_FRAGILE
            violations.append(f"[Fragile] {PRODUCTS[i][0]} in {chromosome[i]} not Z3 → +{PENALTY_FRAGILE}")

    # 5.3 Hazardous products must be in Z5
    for i in HAZARDOUS_IDX:
        if chromosome[i] != "Z5":
            penalty += PENALTY_HAZARDOUS
            violations.append(f"[Hazardous] {PRODUCTS[i][0]} in {chromosome[i]} not Z5 → +{PENALTY_HAZARDOUS}")

    # 5.4 Temperature-sensitive must be in Z4 or Z8
    for i in TEMP_IDX:
        if chromosome[i] not in ("Z4", "Z8"):
            penalty += PENALTY_TEMP
            violations.append(f"[TempCtrl] {PRODUCTS[i][0]} in {chromosome[i]} not Z4/Z8 → +{PENALTY_TEMP}")

    # 5.5 High-demand products should be in Z6
    for i in HIGH_DEMAND_IDX:
        if chromosome[i] != "Z6":
            penalty += PENALTY_DEMAND
            violations.append(f"[Demand] {PRODUCTS[i][0]} in {chromosome[i]} not Z6 → +{PENALTY_DEMAND}")

    # 5.6 Heavy items (>40 kg) must be in Z1
    for i in HEAVY_IDX:
        if chromosome[i] != "Z1":
            penalty += PENALTY_HEAVY
            violations.append(f"[Heavy] {PRODUCTS[i][0]} in {chromosome[i]} not Z1 → +{PENALTY_HEAVY}")

    # 5.7 Same-category compatibility (pairs in different zones)
    # Group by category
    from collections import defaultdict
    category_map = defaultdict(list)
    for i, p in enumerate(PRODUCTS):
        category_map[p[2]].append(i)
    for cat, idxs in category_map.items():
        for a, b in combinations(idxs, 2):
            if chromosome[a] != chromosome[b]:
                penalty += PENALTY_COMPAT
                violations.append(
                    f"[Compat] {PRODUCTS[a][0]} ({chromosome[a]}) and "
                    f"{PRODUCTS[b][0]} ({chromosome[b]}) in diff zones → +{PENALTY_COMPAT}"
                )

    # 5.8 Food-Chemical incompatibility in same zone
    for fi in FOOD_IDX:
        for ci in CHEMICAL_IDX:
            if chromosome[fi] == chromosome[ci]:
                penalty += PENALTY_FOOD_CHEM
                violations.append(
                    f"[FoodChem] {PRODUCTS[fi][0]} & {PRODUCTS[ci][0]} both in {chromosome[fi]} → +{PENALTY_FOOD_CHEM}"
                )

    # 5.9 Z8 non-eligible products
    for prod_idx, zone in enumerate(chromosome):
        if zone == "Z8" and prod_idx not in Z8_ELIGIBLE_IDX:
            penalty += PENALTY_Z8_NONELIG
            violations.append(f"[Z8] {PRODUCTS[prod_idx][0]} not eligible for Z8 → +{PENALTY_Z8_NONELIG}")

    return penalty, violations


# ─────────────────────────────────────────────
# 6. GENETIC ALGORITHM
# ─────────────────────────────────────────────

def random_chromosome():
    """Generate a random chromosome (list of zone assignments)."""
    return [random.choice(ZONE_KEYS) for _ in range(NUM_PRODUCTS)]

def initial_population(pop_size):
    return [random_chromosome() for _ in range(pop_size)]

# --- Selection: Tournament Selection ---
def tournament_selection(population, fitnesses, k=3):
    """
    Picks k individuals at random, returns the one with lowest (best) fitness.
    Tournament selection balances selection pressure vs diversity better than
    roulette wheel (which can be dominated by very fit individuals early on)
    and is simpler/faster than rank selection for large populations.
    """
    selected = random.sample(range(len(population)), k)
    best = min(selected, key=lambda i: fitnesses[i])
    return population[best]

# --- Crossover: Two-Point Crossover ---
def two_point_crossover(parent1, parent2):
    """
    Two-point crossover preserves larger contiguous blocks from each parent,
    giving better exploration of the search space than single-point crossover.
    """
    size = len(parent1)
    p1, p2 = sorted(random.sample(range(size), 2))
    child1 = parent1[:p1] + parent2[p1:p2] + parent1[p2:]
    child2 = parent2[:p1] + parent1[p1:p2] + parent2[p2:]
    return child1, child2

# --- Mutation: Random Gene Reassignment ---
def mutate(chromosome, mutation_rate=0.05):
    """
    Each gene independently mutated with probability mutation_rate.
    Reassigns a random zone to the product at that position.
    """
    chrom = chromosome[:]
    for i in range(len(chrom)):
        if random.random() < mutation_rate:
            chrom[i] = random.choice(ZONE_KEYS)
    return chrom

# --- Main GA Loop ---
def genetic_algorithm(
    pop_size=100,
    max_generations=300,
    mutation_rate=0.05,
    elitism=2,
    tournament_k=3,
    verbose=True
):
    """
    Run the Genetic Algorithm.

    Parameters:
        pop_size        : number of individuals in the population (50–200)
        max_generations : maximum number of generations (100–500)
        mutation_rate   : probability of mutating each gene (0.01–0.1)
        elitism         : number of top individuals carried forward unchanged
        tournament_k    : tournament size for selection
        verbose         : print progress every 50 generations
    """
    population = initial_population(pop_size)
    best_chromosome = None
    best_fitness = float("inf")
    best_violations = []
    best_gen = 0

    for gen in range(1, max_generations + 1):
        # Evaluate fitness for the whole population
        eval_results = [fitness(chrom) for chrom in population]
        fitnesses = [r[0] for r in eval_results]

        # Track overall best
        gen_best_idx = min(range(pop_size), key=lambda i: fitnesses[i])
        if fitnesses[gen_best_idx] < best_fitness:
            best_fitness = fitnesses[gen_best_idx]
            best_chromosome = population[gen_best_idx][:]
            best_violations = eval_results[gen_best_idx][1]
            best_gen = gen

        if verbose and (gen % 50 == 0 or gen == 1 or best_fitness == 0):
            print(f"  Generation {gen:>4} | Best fitness so far: {best_fitness}")

        # Termination: perfect solution found
        if best_fitness == 0:
            break

        # Build next generation
        # Elitism: carry top individuals unchanged
        sorted_pop = sorted(zip(fitnesses, population), key=lambda x: x[0])
        next_gen = [ind for _, ind in sorted_pop[:elitism]]

        # Fill the rest via selection + crossover + mutation
        while len(next_gen) < pop_size:
            p1 = tournament_selection(population, fitnesses, tournament_k)
            p2 = tournament_selection(population, fitnesses, tournament_k)
            c1, c2 = two_point_crossover(p1, p2)
            c1 = mutate(c1, mutation_rate)
            c2 = mutate(c2, mutation_rate)
            next_gen.append(c1)
            if len(next_gen) < pop_size:
                next_gen.append(c2)

        population = next_gen

    return best_chromosome, best_fitness, best_gen, best_violations


# ─────────────────────────────────────────────
# 7. OUTPUT
# ─────────────────────────────────────────────

def print_results(chromosome, best_fitness, generations_run, violations):
    print("\n" + "=" * 60)
    print("           OPTIMAL STORAGE PLAN")
    print("=" * 60)

    zone_products = {z: [] for z in ZONE_KEYS}
    for prod_idx, zone in enumerate(chromosome):
        zone_products[zone].append(PRODUCTS[prod_idx][0])

    for zone in ZONE_KEYS:
        zone_name = ZONES[zone]["name"]
        contents = zone_products[zone]
        if contents:
            print(f"{zone} ({zone_name}):")
            for p in contents:
                print(f"    - {p}")
        else:
            print(f"{zone} ({zone_name}): (empty)")

    print("\n" + "=" * 60)
    print(f"Best Fitness Score : {best_fitness}")
    print(f"Generations Run    : {generations_run}")
    print("=" * 60)

    if violations:
        print(f"\nRemaining Constraint Violations ({len(violations)}):")
        for v in violations:
            print(f"  ✗ {v}")
    else:
        print("\n✓ All constraints satisfied! Perfect solution found.")


# ─────────────────────────────────────────────
# 8. BONUS: EXCEL EXPORT
# ─────────────────────────────────────────────

def export_to_excel(chromosome, filename="warehouse_storage_plan.xlsx"):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Storage Plan"

        # Zone colour palette
        zone_colors = {
            "Z1": "FFD700", "Z2": "ADD8E6", "Z3": "90EE90",
            "Z4": "87CEEB", "Z5": "FF6347", "Z6": "FFA500",
            "Z7": "D3D3D3", "Z8": "B0E0E6",
        }

        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Header row
        headers = ["Zone", "Zone Name", "Capacity (kg)", "Product", "Weight (kg)",
                   "Category", "Fragile", "Hazardous", "Temp Ctrl", "Demand"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2F4F4F")
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        # Data rows
        zone_products = {z: [] for z in ZONE_KEYS}
        for prod_idx, zone in enumerate(chromosome):
            zone_products[zone].append(prod_idx)

        for zone in ZONE_KEYS:
            idxs = zone_products[zone]
            if not idxs:
                row = [zone, ZONES[zone]["name"], ZONES[zone]["capacity"],
                       "(empty)", "", "", "", "", "", ""]
                ws.append(row)
                for cell in ws[ws.max_row]:
                    cell.fill = PatternFill("solid", fgColor=zone_colors[zone])
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center")
            else:
                for prod_idx in idxs:
                    p = PRODUCTS[prod_idx]
                    row = [
                        zone, ZONES[zone]["name"], ZONES[zone]["capacity"],
                        p[0], p[1], p[2],
                        "Yes" if p[3] else "No",
                        "Yes" if p[4] else "No",
                        "Yes" if p[5] else "No",
                        p[6],
                    ]
                    ws.append(row)
                    for cell in ws[ws.max_row]:
                        cell.fill = PatternFill("solid", fgColor=zone_colors[zone])
                        cell.border = border
                        cell.alignment = Alignment(horizontal="center")

        # Auto-width columns
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

        wb.save(filename)
        print(f"\n✓ Excel report saved to '{filename}'")
    except ImportError:
        print("\n[Bonus] openpyxl not installed. Run:  pip install openpyxl")


# ─────────────────────────────────────────────
# 9. ENTRY POINT
# ─────────────────────────────────────────────

if __name__ == "__main__":
    # ── Configurable parameters ──────────────────────────
    POPULATION_SIZE  = 150   # 50–200 (larger → better diversity)
    MAX_GENERATIONS  = 400   # 100–500
    MUTATION_RATE    = 0.05  # 1–10 % per gene
    ELITISM_COUNT    = 3     # top individuals preserved unchanged
    TOURNAMENT_SIZE  = 4     # tournament selection pressure
    EXPORT_EXCEL     = True  # set False to skip bonus Excel output
    # ─────────────────────────────────────────────────────

    print("Warehouse Storage Optimization — Genetic Algorithm")
    print(f"Population: {POPULATION_SIZE}  |  Max Generations: {MAX_GENERATIONS}  "
          f"|  Mutation Rate: {MUTATION_RATE}")
    print("-" * 60)

    best_chrom, best_fit, gens_run, violations = genetic_algorithm(
        pop_size=POPULATION_SIZE,
        max_generations=MAX_GENERATIONS,
        mutation_rate=MUTATION_RATE,
        elitism=ELITISM_COUNT,
        tournament_k=TOURNAMENT_SIZE,
        verbose=True,
    )

    print_results(best_chrom, best_fit, gens_run, violations)

    if EXPORT_EXCEL:
        export_to_excel(best_chrom)
